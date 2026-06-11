"""
提取 Excel 超链接脚本
从【付鹏的财经世界】文章目录索引.xlsx 中提取真实 URL
输出: final_links_with_url.csv
"""

import openpyxl
import csv
import os

BASE = "D:/Quan_Strategy/06 FICC与宏观逻辑"
XLSX_PATH = f"{BASE}/【付鹏的财经世界】文章目录索引.xlsx"
CSV_PATH = f"{BASE}/final_links_with_url.csv"

def extract_links():
    print(f"读取文件: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    results = []
    no_url_count = 0
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n处理 Sheet: {sheet_name} ({ws.max_row} 行 x {ws.max_column} 列)")

        # 跳过表头（第1行）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            total += 1
            date_val = ""
            title_val = ""
            url_val = ""

            # 尝试从各列提取数据
            for col_idx, cell in enumerate(row, start=1):
                val = str(cell.value).strip() if cell.value is not None else ""
                hyperlink = cell.hyperlink.target if cell.hyperlink else None

                # 日期列（第1列）
                if col_idx == 1:
                    date_val = val

                # 标题列（第2列）或含超链接的列
                if col_idx == 2:
                    title_val = val
                    if hyperlink and hyperlink.startswith("http"):
                        url_val = hyperlink

                # 如果其他列有 URL 且还没找到
                if not url_val and hyperlink and hyperlink.startswith("http"):
                    url_val = hyperlink
                    if not title_val:
                        title_val = val

            if url_val:
                results.append({
                    "日期": date_val,
                    "标题": title_val,
                    "真实URL网址": url_val
                })
            else:
                no_url_count += 1
                if no_url_count <= 5:
                    print(f"  [无URL] 第{row_idx}行: {date_val} | {title_val[:40]}")

    print(f"\n===== 提取结果 =====")
    print(f"总行数:     {total}")
    print(f"有URL行数:  {len(results)}")
    print(f"无URL行数:  {no_url_count}")

    # 写入 CSV
    with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["日期", "标题", "真实URL网址"])
        writer.writeheader()
        writer.writerows(results)

    print(f"\n✅ 已保存: {CSV_PATH}")

    # 校验：确认每行都有 http 开头的 URL
    bad = [r for r in results if not r["真实URL网址"].startswith("http")]
    if bad:
        print(f"⚠️ 警告：{len(bad)} 行 URL 不以 http 开头！")
        for b in bad[:3]:
            print(f"  {b}")
    else:
        print(f"✅ 校验通过：全部 {len(results)} 行均有 http 开头的 URL")

    return results

if __name__ == "__main__":
    extract_links()
